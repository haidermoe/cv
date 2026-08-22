import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import pandas as pd
import os
import sys
import csv
import tempfile
import shutil

def _with_temp_copy(filepath, func):
    if not os.path.exists(filepath):
        return func(filepath)
    ext = os.path.splitext(filepath)[1]
    fd, temp_path = tempfile.mkstemp(suffix=ext)
    os.close(fd)
    try:
        shutil.copy2(filepath, temp_path)
    except PermissionError:
        try:
            os.remove(temp_path)
        except Exception:
            pass
        return func(filepath)
    except Exception:
        pass
    
    try:
        return func(temp_path)
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass

def get_sheet_names(filepath):
    def _do(fp):
        if filepath.lower().endswith('.csv'):
            return [os.path.basename(filepath)]
        elif filepath.lower().endswith('.xls'):
            xl = pd.ExcelFile(fp)
            return xl.sheet_names
        else:
            wb = openpyxl.load_workbook(fp, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
    return _with_temp_copy(filepath, _do)

def read_sheet_data(filepath, sheet_name):
    """
    Reads an Excel sheet and returns a tuple: (data_dict, actual_max_r, actual_max_c)
    data_dict maps (row, col) 1-indexed to its string value or formula.
    """
    if not os.path.exists(filepath):
        return {}, 0, 0
    def _do(fp):
        data = {}
        actual_max_r = 0
        actual_max_c = 0
        if filepath.lower().endswith('.csv'):
            with open(fp, 'r', encoding='utf-8-sig', errors='ignore') as f:
                reader = csv.reader(f)
                for r_idx, row in enumerate(reader, start=1):
                    for c_idx, val in enumerate(row, start=1):
                        if val is not None and str(val).strip() != "":
                            data[(r_idx, c_idx)] = str(val)
                            if r_idx > actual_max_r: actual_max_r = r_idx
                            if c_idx > actual_max_c: actual_max_c = c_idx
        elif filepath.lower().endswith('.xls'):
            # For .xls, we read using pandas (values only, formulas are lost)
            df = pd.read_excel(fp, sheet_name=sheet_name, header=None)
            max_r = len(df)
            max_c = len(df.columns) if max_r > 0 else 0
            for r in range(max_r):
                for c in range(max_c):
                    val = df.iat[r, c]
                    if pd.notna(val) and str(val).strip() != "":
                        data[(r + 1, c + 1)] = str(val)
                        if (r + 1) > actual_max_r: actual_max_r = r + 1
                        if (c + 1) > actual_max_c: actual_max_c = c + 1
        else:
            # For .xlsx, read using openpyxl. Set data_only=True to get calculated values instead of formulas
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for c_idx, val in enumerate(row, start=1):
                        if val is not None and str(val).strip() != "":
                            data[(r_idx, c_idx)] = str(val)
                            if r_idx > actual_max_r: actual_max_r = r_idx
                            if c_idx > actual_max_c: actual_max_c = c_idx
            wb.close()
        return data, actual_max_r, actual_max_c
    return _with_temp_copy(filepath, _do)

import string

def col2num(col_str):
    num = 0
    for c in col_str.upper():
        if c in string.ascii_uppercase:
            num = num * 26 + (ord(c) - ord('A')) + 1
    return num

def get_headers(filepath, sheet_name):
    """
    Reads the first row of a sheet and returns a list of formatted header strings:
    ['A (Header1)', 'B (Header2)', ...]
    """
    if not os.path.exists(filepath):
        return []
    def _do(fp):
        headers = []
        if filepath.lower().endswith('.csv'):
            with open(fp, 'r', encoding='utf-8-sig', errors='ignore') as f:
                reader = csv.reader(f)
                try:
                    row = next(reader)
                    for i, val in enumerate(row):
                        col_letter = openpyxl.utils.get_column_letter(i + 1)
                        v = str(val).strip() if val is not None else ""
                        headers.append(f"{col_letter} ({v})")
                except StopIteration:
                    pass
        else:
            df = pd.read_excel(fp, sheet_name=sheet_name, header=None, nrows=1)
            if not df.empty:
                for i, val in enumerate(df.iloc[0]):
                    col_letter = openpyxl.utils.get_column_letter(i + 1)
                    v = str(val) if pd.notna(val) else ""
                    headers.append(f"{col_letter} ({v})")
        return headers
    return _with_temp_copy(filepath, _do)
import re
import difflib

def extract_numbers(text):
    return set(re.findall(r'\d+', str(text)))

def smart_similarity(old_name, new_name):
    if not old_name or not new_name:
        return 0.0
    
    s_old = str(old_name).lower()
    s_new = str(new_name).lower()
    
    num_old = extract_numbers(s_old)
    num_new = extract_numbers(s_new)
    
    if num_old != num_new:
        return 0.0  # Strict number matching: if numbers differ, do not match.
        
    import string
    s_old_clean = s_old.translate(str.maketrans('', '', string.punctuation)).strip()
    s_new_clean = s_new.translate(str.maketrans('', '', string.punctuation)).strip()
    
    # 1. Standard ratio
    ratio_standard = difflib.SequenceMatcher(None, s_old_clean, s_new_clean).ratio()
    
    # 2. Token Sort Ratio (handles out of order words)
    sorted_old = " ".join(sorted(s_old_clean.split()))
    sorted_new = " ".join(sorted(s_new_clean.split()))
    ratio_sort = difflib.SequenceMatcher(None, sorted_old, sorted_new).ratio()
    
    # 3. Token Set Ratio (handles extra words, e.g., "CREED")
    tokens_old = set(s_old_clean.split())
    tokens_new = set(s_new_clean.split())
    
    intersection = tokens_old.intersection(tokens_new)
    rest_old = tokens_old - intersection
    rest_new = tokens_new - intersection
    
    t0 = " ".join(sorted(intersection))
    t1 = (t0 + " " + " ".join(sorted(rest_old))).strip()
    t2 = (t0 + " " + " ".join(sorted(rest_new))).strip()
    
    ratio_set = 0.0
    if t1 and t2:
        ratio_set = max(
            difflib.SequenceMatcher(None, t0, t1).ratio() if t0 else 0,
            difflib.SequenceMatcher(None, t0, t2).ratio() if t0 else 0,
            difflib.SequenceMatcher(None, t1, t2).ratio()
        )
        
    ratio = max(ratio_standard, ratio_sort, ratio_set)
    return ratio

import urllib.request
import urllib.error
import json

def run_ai_standardizer(items, api_key, cache, progress_callback=None):
    if not api_key or not items:
        return {}
        
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    
    unique_names = []
    seen_names = set()
    for item_key, item_name in items:
        if item_name not in cache and item_name not in seen_names:
            seen_names.add(item_name)
            unique_names.append(item_name)
            
    chunk_size = 50
    total_chunks = (len(unique_names) + chunk_size - 1) // chunk_size
    
    import time
    for i in range(0, len(unique_names), chunk_size):
        if progress_callback:
            progress_callback(f"توحيد الأسماء بالذكاء الاصطناعي... ({i//chunk_size + 1}/{total_chunks})", 0.5)
            
        chunk_names = unique_names[i:i+chunk_size]
        text = ""
        for idx, name in enumerate(chunk_names):
            text += f'- ID: "{idx}" | Name: "{name}"\n'
            
        prompt = f"""You are a video game and electronics store assistant.
Your task is to take a list of raw product names and convert each to its "Global Standard Name" (e.g., 'gta v' -> 'Grand Theft Auto 5', 'Playstation 5' -> 'Sony PlayStation 5').
CRITICAL INSTRUCTIONS:
- Pay close attention to abbreviations! 'gta v' = 'Grand Theft Auto 5', 're' = 'Resident Evil', 'rdr2' = 'Red Dead Redemption 2', 'cod' = 'Call of Duty', 'fc24' = 'EA Sports FC 24'.
- Pay close attention to versions and editions! 'Remake', 'Deluxe', 'PS4', 'PS5'. Keep them in the standard name (e.g., 'Resident Evil 4 Remake - PS5').
- NEVER use Roman numerals (I, II, III, IV, V, etc). Convert ALL Roman numerals to Arabic digits (1, 2, 3, 4, 5). e.g., 'Grand Theft Auto V' -> 'Grand Theft Auto 5'.
- Do not translate to Arabic, keep standard English names for games/consoles unless it's an Arabic specific item.

List of items:
{text}

Respond ONLY with a valid JSON array of objects mapping the IDs to their standard names.
Format:
[
  {{"id": "ID here", "std_name": "Standard Name here"}}
]
"""
        data = {
            "contents": [{"parts":[{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.0,
                "responseMimeType": "application/json"
            }
        }
        
        req = urllib.request.Request(url, data=json.dumps(data).encode('utf-8'), headers={'Content-Type':'application/json'})
        
        retries = 3
        for attempt in range(retries):
            try:
                response = urllib.request.urlopen(req)
                response_text = response.read().decode('utf-8')
                result_json = json.loads(response_text)
                
                candidates = result_json.get('candidates', [])
                if candidates:
                    content_text = candidates[0].get('content', {}).get('parts', [{}])[0].get('text', '[]')
                    try:
                        content_text = content_text.strip()
                        if content_text.startswith("```json"):
                            content_text = content_text[7:]
                        if content_text.startswith("```"):
                            content_text = content_text[3:]
                        if content_text.endswith("```"):
                            content_text = content_text[:-3]
                        content_text = content_text.strip()
                        parsed = json.loads(content_text)
                        for item in parsed:
                            if 'id' in item and 'std_name' in item:
                                idx_str = str(item['id'])
                                if idx_str.isdigit() and int(idx_str) < len(chunk_names):
                                    cache[chunk_names[int(idx_str)]] = str(item['std_name'])
                    except Exception as e:
                        print("Failed to parse JSON from AI:", e)
                
                # Respect rate limit of 15 RPM (1 request every 4 seconds)
                time.sleep(4)
                break # Success, break retry loop
            except urllib.error.HTTPError as e:
                print(f"AI API HTTP Error on attempt {attempt+1}:", e.code)
                if e.code == 429: # Too many requests
                    print("Rate limit hit, waiting 20 seconds...")
                    time.sleep(20)
                else:
                    time.sleep(5)
            except Exception as e:
                print(f"AI API Error on attempt {attempt+1}:", e)
                time.sleep(5)
                
    # Now build the final results from the cache
    results = {}
    for item_key, item_name in items:
        if item_name in cache:
            results[item_key] = cache[item_name]
            
    return results

def clean_key_str(s, ignore_punct=True):
    if s is None:
        return ""
    s = str(s).strip()
    if s.startswith('\ufeff'):
        s = s[1:]
    s = s.replace('\xa0', ' ').replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
    s = re.sub(r'[\u200b\u200c\u200d\u200e\u200f\u202a-\u202e\u0640]', '', s)
    if s.endswith('.0') and s[:-2].replace('.', '', 1).isdigit():
        s = s[:-2]
    s = s.strip().lower()
    s = re.sub(r'[أإآ]', 'ا', s)
    s = re.sub(r'ة', 'ه', s)
    s = re.sub(r'ى', 'ي', s)
    if ignore_punct:
        s = re.sub(r'\W+|_', '', s)
    else:
        s = re.sub(r'\s+', ' ', s)
        s = s.strip(" \t\n\r./-_")
    return s

def lookup_ref_mapping(raw_k, ref_mapping):
    if not raw_k or not ref_mapping:
        return None
    raw_str = str(raw_k).replace('\xa0', ' ').strip()
    k_lower = raw_str.lower()
    k_clean_punct = clean_key_str(raw_str, True)
    k_clean_space = clean_key_str(raw_str, False)
    
    if k_lower in ref_mapping:
        return ref_mapping[k_lower]
    if k_clean_punct in ref_mapping:
        return ref_mapping[k_clean_punct]
    if k_clean_space in ref_mapping:
        return ref_mapping[k_clean_space]
    return None

def compare_excel_files(old_path, new_path, output_path, key_col_old="", key_col_new="", target_sheet="", comp_col_old="", comp_col_new="", ignore_punct=True, similarity_threshold=70, use_ai=False, ai_api_key="", mode="compare", new_col_name="", filter_keywords="", ref_filepath="", progress_callback=None):

    """
    Compares two Excel files and generates a color-coded output report.
    Also generates a '_Changes' sheet containing only the modified rows.
    """
    # Define styles
    fill_modified = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
    fill_new = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")       # Light Green
    fill_deleted = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # Light Red

    key_col_idx_old = col2num(key_col_old.split()[0]) if key_col_old and key_col_old != "Row-by-Row" else None
    key_col_idx_new = col2num(key_col_new.split()[0]) if key_col_new and key_col_new != "Row-by-Row" else None
    
    comp_col_idx_old = col2num(comp_col_old.split()[0]) if comp_col_old and comp_col_old != "Compare All Columns" else None
    comp_col_idx_new = col2num(comp_col_new.split()[0]) if comp_col_new and comp_col_new != "Compare All Columns" else None

    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active) # Remove default sheet
    
    old_sheets = get_sheet_names(old_path)
    new_sheets = get_sheet_names(new_path)
    
    is_old_csv = old_path.lower().endswith('.csv')
    is_new_csv = new_path.lower().endswith('.csv')
    
    if is_old_csv and not is_new_csv:
        common_sheets = new_sheets
    elif is_new_csv and not is_old_csv:
        common_sheets = old_sheets
    elif is_old_csv and is_new_csv:
        common_sheets = ["CSV Data"]
    else:
        # Maintain order, avoid duplicates
        common_sheets = list(dict.fromkeys([s for s in old_sheets if s in new_sheets]))
        
    if target_sheet and target_sheet != "Compare All Sheets":
        if target_sheet in common_sheets:
            all_sheets = [target_sheet]
        else:
            raise ValueError(f"Sheet '{target_sheet}' not found or no common sheets available.")
    else:
        all_sheets = common_sheets

    total_sheets = len(all_sheets)
    
    def create_unique_sheet(wb, desired_title):
        base = desired_title[:31]
        if base not in wb.sheetnames:
            return wb.create_sheet(title=base)
        suffix = 1
        while True:
            suffix_str = f"_{suffix}"
            max_len = 31 - len(suffix_str)
            candidate = f"{desired_title[:max_len]}{suffix_str}"
            if candidate not in wb.sheetnames:
                return wb.create_sheet(title=candidate)
            suffix += 1

    for i, sheet_name in enumerate(all_sheets):
        if progress_callback:
            progress_callback(f"Processing sheet: {sheet_name} ({i+1}/{total_sheets})", (i) / total_sheets)
            
        out_sheet = create_unique_sheet(out_wb, sheet_name)
        
        # Create categorized sheets
        base_title = sheet_name[:20]
        mod_sheet = create_unique_sheet(out_wb, f"{base_title}_Mod")
        new_sheet = create_unique_sheet(out_wb, f"{base_title}_New")
        del_sheet = create_unique_sheet(out_wb, f"{base_title}_Del")
        oos_sheet = create_unique_sheet(out_wb, f"{base_title}_Out_Of_Stock")
        sug_sheet = create_unique_sheet(out_wb, f"{base_title}_Suggestions")
        
        old_data, old_max_r, old_max_c = read_sheet_data(old_path, sheet_name)
        new_data, new_max_r, new_max_c = read_sheet_data(new_path, sheet_name)
        
        max_r_total = max(old_max_r, new_max_r)
        
        # If both sheets are completely empty
        if max_r_total == 0 and old_max_c == 0 and new_max_c == 0:
            continue

        def matches_keywords(raw_k, kw_str):
            if not kw_str:
                return True
            if not raw_k:
                return False
            kws = [k.strip().lower() for k in kw_str.split(',') if k.strip()]
            if not kws:
                return True
            raw_l = str(raw_k).lower()
            for kw in kws:
                if kw in raw_l:
                    return True
            return False
            
        # --- Reference File Processing ---
        ref_mapping = {}
        if ref_filepath and os.path.exists(ref_filepath):
            try:
                ref_rows = []
                if ref_filepath.lower().endswith('.csv'):
                    with open(ref_filepath, 'r', encoding='utf-8-sig', errors='ignore') as f:
                        reader = csv.reader(f)
                        ref_rows = [row for row in reader]
                elif ref_filepath.lower().endswith('.xls'):
                    xls_sheets = get_sheet_names(ref_filepath)
                    for sh in xls_sheets:
                        df = pd.read_excel(ref_filepath, sheet_name=sh, header=None)
                        for _, row in df.iterrows():
                            ref_rows.append([str(v) if pd.notna(v) else "" for v in row])
                else:
                    ref_wb = openpyxl.load_workbook(ref_filepath, data_only=True, read_only=True)
                    for sheet_name in ref_wb.sheetnames:
                        ws = ref_wb[sheet_name]
                        for row in ws.iter_rows(values_only=True):
                            ref_rows.append(list(row))
                    ref_wb.close()

                for r_idx, row in enumerate(ref_rows, 1):
                    if not row:
                        continue
                    vals = [str(v).replace('\xa0', ' ').strip() for v in row if v is not None and str(v).strip() != ""]
                    if r_idx == 1:
                        row_str = " ".join(vals).lower()
                        if any(h in row_str for h in ['name', 'sku', 'اسم', 'موقع', 'مخزن', 'item', 'ref', 'كود', 'اصلي', 'القديم', 'المحدث']):
                            continue
                    if len(vals) >= 2:
                        target_std = vals[2] if len(vals) >= 3 else vals[1]
                        for v in vals:
                            if not v: continue
                            ref_mapping[v.lower()] = target_std
                            ref_mapping[clean_key_str(v, True)] = target_std
                            ref_mapping[clean_key_str(v, False)] = target_std
            except Exception as e:
                print("Failed to load reference file:", e)

        # --- AI Pre-processing ---
        ai_cache = {}
        cache_path = ""
        std_mapping_old = {}
        std_mapping_new = {}
        
        if key_col_idx_old and key_col_idx_new and use_ai and ai_api_key:
            if getattr(sys, 'frozen', False):
                app_dir = os.path.dirname(sys.executable)
            else:
                app_dir = os.path.dirname(os.path.abspath(__file__))
            cache_path = os.path.join(app_dir, "ai_cache.json")
            if os.path.exists(cache_path):
                try:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        ai_cache = json.load(f)
                except:
                    pass
                    
            # Gather all names
            old_list = []
            for r in range(1, old_max_r + 1):
                raw_key = old_data.get((r, key_col_idx_old))
                if raw_key and matches_keywords(raw_key, filter_keywords): 
                    old_list.append((str(r), raw_key))
                
            new_list = []
            for r in range(1, new_max_r + 1):
                raw_key = new_data.get((r, key_col_idx_new))
                if raw_key and matches_keywords(raw_key, filter_keywords): 
                    new_list.append((str(r), raw_key))
                
            std_mapping_old = run_ai_standardizer(old_list, ai_api_key, ai_cache, progress_callback)
            std_mapping_new = run_ai_standardizer(new_list, ai_api_key, ai_cache, progress_callback)
            
            if cache_path:
                try:
                    with open(cache_path, "w", encoding="utf-8") as f:
                        json.dump(ai_cache, f, ensure_ascii=False, indent=2)
                except:
                    pass

        if mode == "pull":
            out_wb.remove(mod_sheet)
            out_wb.remove(new_sheet)
            out_wb.remove(del_sheet)
            out_wb.remove(oos_sheet)
            out_wb.remove(sug_sheet)
            
            def normalize_key(raw_k, row_idx=None, is_old=False):
                if raw_k is None:
                    return ""
                
                ref_val = lookup_ref_mapping(raw_k, ref_mapping)
                if ref_val:
                    k = ref_val
                else:
                    k = raw_k
                    if use_ai and ai_api_key and row_idx is not None:
                        key_id = str(row_idx)
                        if is_old and key_id in std_mapping_old and std_mapping_old[key_id]:
                            k = std_mapping_old[key_id]
                        elif not is_old and key_id in std_mapping_new and std_mapping_new[key_id]:
                            k = std_mapping_new[key_id]
                        
                s = str(k).lower()
                if ignore_punct:
                    s = re.sub(r'\W+|_', '', s)
                else:
                    s = s.replace('\\', '/')
                    s = s.strip(" \t\n\r./-_")
                return s
                
            old_by_key = {}
            old_keys_order = []
            row_to_key_old = {}
            base_key_to_row_old = {}
            for r in range(1, old_max_r + 1):
                raw_key = old_data.get((r, key_col_idx_old))
                ref_k = lookup_ref_mapping(raw_key, ref_mapping) or raw_key
                if not matches_keywords(raw_key, filter_keywords) and not matches_keywords(ref_k, filter_keywords): continue
                norm_k = normalize_key(raw_key, r, True)
                key = norm_k
                if not key or key == "":
                    key = f"__ROW__{r}"
                if key in old_by_key and not str(key).startswith("__ROW__"):
                    key = f"{key}__DUP_{r}"
                row_dict = {c: old_data.get((r, c)) for c in range(1, old_max_c + 1)}
                if use_ai and ai_api_key and str(r) in std_mapping_old and std_mapping_old[str(r)]:
                    row_dict['__std_name__'] = std_mapping_old[str(r)]
                old_by_key[key] = row_dict
                old_keys_order.append(key)
                row_to_key_old[r] = key
                base_key_to_row_old[r] = norm_k
                
            new_by_key = {}
            for r in range(1, new_max_r + 1):
                raw_key = new_data.get((r, key_col_idx_new))
                ref_k = lookup_ref_mapping(raw_key, ref_mapping) or raw_key
                if not matches_keywords(raw_key, filter_keywords) and not matches_keywords(ref_k, filter_keywords): continue
                key = normalize_key(raw_key, r, False)
                if not key or key == "":
                    key = f"__ROW__{r}"
                if key in new_by_key and not str(key).startswith("__ROW__"):
                    key = f"{key}__DUP_{r}"
                row_dict = {c: new_data.get((r, c)) for c in range(1, new_max_c + 1)}
                if use_ai and ai_api_key and str(r) in std_mapping_new and std_mapping_new[str(r)]:
                    row_dict['__std_name__'] = std_mapping_new[str(r)]
                new_by_key[key] = row_dict
                
            unmatched_old_keys = [k for k in old_keys_order if k not in new_by_key]
            unmatched_new_keys = [k for k in new_by_key.keys() if k not in old_by_key]
            
            suggestions_to_review = []
            matched_by_fuzzy = {}
            
            ai_cache = {}
            for ok in unmatched_old_keys:
                if ok in matched_by_fuzzy:
                    continue
                old_row = old_by_key[ok]
                old_name = old_row.get('__std_name__') or old_row.get(key_col_idx_old) or ok
                best_match = None
                best_ratio = 0.0
                for nk in unmatched_new_keys:
                    new_row = new_by_key[nk]
                    new_name = new_row.get('__std_name__') or new_row.get(key_col_idx_new) or nk
                    ratio = smart_similarity(old_name, new_name)
                    if ratio > best_ratio:
                        best_ratio = ratio
                        best_match = nk
                if best_ratio >= (similarity_threshold / 100.0) and best_match:
                    matched_by_fuzzy[ok] = best_match
                    
            new_col_idx = old_max_c + 1
            fill_increase = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Light Green
            fill_decrease = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
            header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")   # Soft Blue

            for r in range(1, old_max_r + 1):
                key = row_to_key_old.get(r)
                base_key = base_key_to_row_old.get(r)
                
                for c in range(1, old_max_c + 1):
                    out_sheet.cell(row=r, column=c).value = old_data.get((r, c))
                    
                if r == 1:
                    out_sheet.cell(row=r, column=new_col_idx).value = new_col_name
                    out_sheet.cell(row=r, column=new_col_idx).fill = header_fill
                else:
                    target_nk = None
                    if base_key and base_key in new_by_key:
                        target_nk = base_key
                    elif key and key in new_by_key:
                        target_nk = key
                    elif key and key in matched_by_fuzzy:
                        target_nk = matched_by_fuzzy[key]
                    elif base_key and base_key in matched_by_fuzzy:
                        target_nk = matched_by_fuzzy[base_key]
                        
                    old_val_raw = old_data.get((r, comp_col_idx_old)) if comp_col_idx_old else None
                    try:
                        old_num = float(str(old_val_raw).strip()) if old_val_raw is not None and str(old_val_raw).strip() != "" else 0.0
                    except:
                        old_num = 0.0

                    if target_nk and target_nk in new_by_key:
                        pulled_val = new_by_key[target_nk].get(comp_col_idx_new)
                        if pulled_val is not None and str(pulled_val).strip() != "":
                            out_sheet.cell(row=r, column=new_col_idx).value = pulled_val
                            try:
                                pulled_num = float(str(pulled_val).strip())
                            except:
                                pulled_num = 0.0
                        else:
                            out_sheet.cell(row=r, column=new_col_idx).value = 0
                            pulled_num = 0.0

                        if pulled_num > old_num:
                            out_sheet.cell(row=r, column=new_col_idx).fill = fill_increase
                        elif pulled_num < old_num:
                            out_sheet.cell(row=r, column=new_col_idx).fill = fill_decrease
                    else:
                        out_sheet.cell(row=r, column=new_col_idx).value = 0
                        if old_num > 0:
                            out_sheet.cell(row=r, column=new_col_idx).fill = fill_decrease
            continue
            
        mod_r = 1
        new_r = 1
        del_r = 1
        oos_r = 1

        def process_row(out_r, old_row, new_row, is_new, is_deleted):
            nonlocal mod_r, new_r, del_r, oos_r
            row_has_changes = False
            cells = []
            
            old_mod_vals = []
            new_mod_vals = []
            
            is_out_of_stock = False
            if is_deleted:
                is_out_of_stock = True
            elif not is_new and comp_col_idx_new is not None:
                qty_val = new_row.get(comp_col_idx_new) if new_row else None
                s = str(qty_val).strip().lower()
                if s in ['0', '0.0', 'out of stock', 'none', 'null', '']:
                    is_out_of_stock = True
                    
            sku_val = ""
            if old_row and key_col_idx_old:
                sku_val = old_row.get(key_col_idx_old)
            elif new_row and key_col_idx_new:
                sku_val = new_row.get(key_col_idx_new)
            else:
                sku_val = f"Row {out_r}"
            
            # Determine how many columns to iterate over.
            # If completely new, we iterate over new_max_c so we capture all the new item's columns.
            # Otherwise, we stick to old_max_c to maintain the website file structure.
            iterate_c = new_max_c if is_new else old_max_c
            
            for c in range(1, iterate_c + 1):
                old_val = old_row.get(c) if old_row else None
                
                # Fetch new_val based on mapping logic
                if is_new:
                    # New items are just dumped as is
                    new_val = new_row.get(c) if new_row else None
                elif comp_col_idx_old is not None and comp_col_idx_new is not None:
                    # Specific comparison column is mapped
                    if c == comp_col_idx_old:
                        new_val = new_row.get(comp_col_idx_new) if new_row else None
                    else:
                        new_val = old_val # Don't update other columns
                else:
                    # Default: compare index-to-index
                    new_val = new_row.get(c) if new_row else None
                
                cell_value = None
                cell_fill = None
                
                if is_new:
                    if new_val is not None:
                        cell_value = f"[New]: {new_val}"
                        cell_fill = fill_new
                        row_has_changes = True
                elif is_deleted:
                    if old_val is not None:
                        cell_value = f"[Deleted]: was ({old_val})"
                        cell_fill = fill_deleted
                        row_has_changes = True
                else:
                    # Modified or Unchanged
                    if comp_col_idx_old is not None and c != comp_col_idx_old:
                        # Skip comparison for non-target columns when a specific comparison column is chosen
                        cell_value = new_val
                    else:
                        if old_val == new_val:
                            cell_value = old_val if old_val is not None else new_val
                        elif old_val is None and new_val is not None:
                            cell_value = f"[New]: {new_val}"
                            cell_fill = fill_new
                            row_has_changes = True
                            new_mod_vals.append(str(new_val))
                        elif old_val is not None and new_val is None:
                            cell_value = f"[Deleted]: was ({old_val})"
                            cell_fill = fill_deleted
                            row_has_changes = True
                            old_mod_vals.append(str(old_val))
                        else:
                            cell_value = new_val
                            cell_fill = fill_modified
                            row_has_changes = True
                            old_mod_vals.append(str(old_val))
                            new_mod_vals.append(str(new_val))
                
                cells.append((cell_value, cell_fill))
                
                # Write to main sheet
                out_cell = out_sheet.cell(row=out_r, column=c)
                if cell_value is not None:
                    out_cell.value = cell_value
                if cell_fill is not None:
                    out_cell.fill = cell_fill
                    
            # Handle headers (write to all categorized sheets)
            if out_r == 1:
                for c, (val, fill) in enumerate(cells, start=1):
                    if val is not None:
                        mod_sheet.cell(row=mod_r, column=c).value = val
                        del_sheet.cell(row=del_r, column=c).value = val
                    if fill is not None:
                        mod_sheet.cell(row=mod_r, column=c).fill = fill
                        del_sheet.cell(row=del_r, column=c).fill = fill

                # Headers for _New sheet ALWAYS take 100% exact structure/headers of the Warehouse File (new_data)
                for c in range(1, new_max_c + 1):
                    h_val = new_data.get((1, c))
                    if h_val is not None:
                        new_sheet.cell(row=new_r, column=c).value = h_val
                    new_sheet.cell(row=new_r, column=c).fill = fill_new
                
                # Add extra headers to _Mod
                status_col = len(cells) + 1
                old_val_col = len(cells) + 2
                new_val_col = len(cells) + 3
                
                mod_sheet.cell(row=mod_r, column=status_col).value = "حالة التحديث"
                mod_sheet.cell(row=mod_r, column=old_val_col).value = "القيمة القديمة"
                mod_sheet.cell(row=mod_r, column=new_val_col).value = "القيمة المحدثة"
                
                oos_sheet.cell(row=oos_r, column=1).value = "Item/SKU"
                oos_sheet.cell(row=oos_r, column=2).value = "Status"
                oos_sheet.cell(row=oos_r, column=3).value = "ملاحظات التشابه"
                
                mod_r += 1
                new_r += 1
                del_r += 1
                oos_r += 1
            else:
                status_col = len(cells) + 1
                old_val_col = len(cells) + 2
                new_val_col = len(cells) + 3
                
                # Distribute row to the appropriate sheet
                if is_new:
                    for c in range(1, new_max_c + 1):
                        clean_val = new_row.get(c) if new_row else None
                        ch_cell = new_sheet.cell(row=new_r, column=c)
                        if clean_val is not None:
                            ch_cell.value = clean_val
                        ch_cell.fill = fill_new
                    new_r += 1
                else:
                    # It's an old item (either deleted, modified, or unchanged) -> write to _Mod
                    for c, (val, fill) in enumerate(cells, start=1):
                        ch_cell = mod_sheet.cell(row=mod_r, column=c)
                        if val is not None: ch_cell.value = val
                        if fill is not None: ch_cell.fill = fill
                    
                    if is_deleted:
                        mod_sheet.cell(row=mod_r, column=status_col).value = "Out of stock"
                        # Also put in _Del sheet for convenience
                        for c, (val, fill) in enumerate(cells, start=1):
                            ch_cell = del_sheet.cell(row=del_r, column=c)
                            if val is not None: ch_cell.value = val
                            if fill is not None: ch_cell.fill = fill
                        del_r += 1
                    elif row_has_changes:
                        if is_out_of_stock:
                            mod_sheet.cell(row=mod_r, column=status_col).value = "Out of stock"
                        else:
                            mod_sheet.cell(row=mod_r, column=status_col).value = "تم التحديث"
                            
                        if old_mod_vals:
                            mod_sheet.cell(row=mod_r, column=old_val_col).value = " | ".join(old_mod_vals)
                            mod_sheet.cell(row=mod_r, column=old_val_col).fill = fill_modified
                        if new_mod_vals:
                            mod_sheet.cell(row=mod_r, column=new_val_col).value = " | ".join(new_mod_vals)
                            mod_sheet.cell(row=mod_r, column=new_val_col).fill = fill_modified
                    else:
                        mod_sheet.cell(row=mod_r, column=status_col).value = "لا تحديث"
                        
                    if is_out_of_stock:
                        if not is_deleted:
                            oos_sheet.cell(row=oos_r, column=1).value = sku_val
                            oos_sheet.cell(row=oos_r, column=2).value = "نفذ الرصيد (موجودة)"
                            oos_sheet.cell(row=oos_r, column=3).value = "-"
                            oos_r += 1
                        
                    mod_r += 1
            
        if key_col_idx_old and key_col_idx_new:
            def normalize_key(raw_k, row_idx=None, is_old=False):
                if raw_k is None:
                    return ""
                
                ref_val = lookup_ref_mapping(raw_k, ref_mapping)
                if ref_val:
                    k = ref_val
                else:
                    k = raw_k
                    if use_ai and ai_api_key and row_idx is not None:
                        key_id = str(row_idx)
                        if is_old and key_id in std_mapping_old and std_mapping_old[key_id]:
                            k = std_mapping_old[key_id]
                        elif not is_old and key_id in std_mapping_new and std_mapping_new[key_id]:
                            k = std_mapping_new[key_id]
                        
                return clean_key_str(k, ignore_punct)
                
            old_by_key = {}
            old_keys_order = []
            for r in range(2, old_max_r + 1):
                raw_key = old_data.get((r, key_col_idx_old))
                ref_k = lookup_ref_mapping(raw_key, ref_mapping) or raw_key
                if not matches_keywords(raw_key, filter_keywords) and not matches_keywords(ref_k, filter_keywords): continue
                key = normalize_key(raw_key, r, True)
                
                if not key or key == "":
                    key = f"__ROW__{r}"
                if key in old_by_key and not str(key).startswith("__ROW__"):
                    key = f"{key}__DUP_{r}"
                row_dict = {c: old_data.get((r, c)) for c in range(1, old_max_c + 1)}
                if use_ai and ai_api_key and str(r) in std_mapping_old and std_mapping_old[str(r)]:
                    row_dict['__std_name__'] = std_mapping_old[str(r)]
                old_by_key[key] = row_dict
                old_keys_order.append(key)
                
            new_by_key = {}
            new_keys_order = []
            for r in range(2, new_max_r + 1):
                raw_key = new_data.get((r, key_col_idx_new))
                ref_k = lookup_ref_mapping(raw_key, ref_mapping) or raw_key
                if not matches_keywords(raw_key, filter_keywords) and not matches_keywords(ref_k, filter_keywords): continue
                key = normalize_key(raw_key, r, False)
                
                if not key or key == "":
                    key = f"__ROW__{r}"
                if key in new_by_key and not str(key).startswith("__ROW__"):
                    key = f"{key}__DUP_{r}"
                row_dict = {c: new_data.get((r, c)) for c in range(1, new_max_c + 1)}
                if use_ai and ai_api_key and str(r) in std_mapping_new and std_mapping_new[str(r)]:
                    row_dict['__std_name__'] = std_mapping_new[str(r)]
                new_by_key[key] = row_dict
                new_keys_order.append(key)
                
            out_r = 1
            # Write Header Row first
            header_old_row = {c: old_data.get((1, c)) for c in range(1, old_max_c + 1)}
            header_new_row = {c: new_data.get((1, c)) for c in range(1, new_max_c + 1)}
            process_row(out_r, header_old_row, header_new_row, False, False)
            out_r += 1

            # Process New and Modified rows based on new file order
            for key in new_keys_order:
                new_row = new_by_key[key]
                matched_old_key = key
                if key not in old_by_key and "__DUP_" in key:
                    base_k = key.split("__DUP_")[0]
                    if base_k in old_by_key:
                        matched_old_key = base_k
                if matched_old_key in old_by_key:
                    old_row = old_by_key[matched_old_key]
                    process_row(out_r, old_row, new_row, False, False)
                else:
                    process_row(out_r, None, new_row, True, False)
                out_r += 1
                
            # Process Deleted rows (in old but not in new)
            for key in old_keys_order:
                base_k = key.split("__DUP_")[0] if "__DUP_" in key else key
                if key not in new_by_key and base_k not in new_by_key:
                    old_row = old_by_key[key]
                    process_row(out_r, old_row, None, False, True)
                    out_r += 1
                    
            # --- Smart Suggestions for Deleted Items ---
            unmatched_old_keys = [k for k in old_keys_order if k not in new_by_key]
            unmatched_new_keys = [k for k in new_keys_order if k not in old_by_key]
            
            sug_sheet.cell(row=1, column=1).value = "المادة القديمة (غير متوفرة)"
            sug_sheet.cell(row=1, column=2).value = "المادة المقترحة من المخزن (تشابه بالاسم)"
            sug_sheet.cell(row=1, column=3).value = "نسبة التشابه"
            sug_sheet.cell(row=1, column=4).value = "الرصيد القديم"
            sug_sheet.cell(row=1, column=5).value = "الرصيد المقترح الجديد"
            sug_sheet.cell(row=1, column=6).value = "حالة التشابه (اكتب out للرفض)"
            
            sug_r = 2
            
            suggestions_to_review = []
            
            for ok in unmatched_old_keys:
                old_row = old_by_key[ok]
                old_name = old_row.get('__std_name__') or old_row.get(key_col_idx_old) or ok
                
                best_match = None
                best_ratio = 0.0
                is_ai_approved = False
                
                # Normal fuzzy matching
                for nk in unmatched_new_keys:
                    new_row = new_by_key[nk]
                    new_name = new_row.get('__std_name__') or new_row.get(key_col_idx_new) or nk
                    ratio = smart_similarity(old_name, new_name)
                    if ratio > best_ratio:
                        best_ratio = ratio
                        best_match = (nk, new_name, new_row)
                        
                if best_ratio >= (similarity_threshold / 100.0) and best_match:
                    suggestions_to_review.append((ok, old_name, best_match, best_ratio, is_ai_approved))
                else:
                    oos_sheet.cell(row=oos_r, column=1).value = old_name
                    oos_sheet.cell(row=oos_r, column=2).value = "محذوفة من المخزن"
                    oos_sheet.cell(row=oos_r, column=3).value = "لا يوجد تشابه"
                    oos_r += 1

            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            for ok, old_name, best_match, best_ratio, is_ai_approved in suggestions_to_review:
                oos_sheet.cell(row=oos_r, column=1).value = old_name
                oos_sheet.cell(row=oos_r, column=2).value = "محذوفة من المخزن"
                
                if is_ai_approved:
                    oos_sheet.cell(row=oos_r, column=3).value = "تمت المطابقة بواسطة الذكاء الاصطناعي (AI)"
                else:
                    oos_sheet.cell(row=oos_r, column=3).value = "يوجد تشابه، راجع شيت الاقتراحات (_Suggestions)!"
                oos_r += 1
                
                sug_sheet.cell(row=sug_r, column=1).value = old_name
                sug_sheet.cell(row=sug_r, column=2).value = best_match[1]
                if is_ai_approved:
                    sug_sheet.cell(row=sug_r, column=3).value = "100% (AI)"
                else:
                    sug_sheet.cell(row=sug_r, column=3).value = f"{int(best_ratio * 100)}%"
                
                old_row = old_by_key[ok]
                old_qty = old_row.get(comp_col_idx_old) if comp_col_idx_old else ""
                new_qty = best_match[2].get(comp_col_idx_new) if comp_col_idx_new else ""
                sug_sheet.cell(row=sug_r, column=4).value = old_qty
                sug_sheet.cell(row=sug_r, column=5).value = new_qty
                
                if is_ai_approved:
                    sug_sheet.cell(row=sug_r, column=6).value = "AI_APPROVED"
                    for col in range(1, 7):
                        sug_sheet.cell(row=sug_r, column=col).fill = green_fill
                else:
                    sug_sheet.cell(row=sug_r, column=6).value = ""
                        
                sug_r += 1
                    
        else:
            # --- Fallback: Row-by-Row Comparison ---
            for r in range(1, max_r_total + 1):
                old_row = {c: old_data.get((r, c)) for c in range(1, old_max_c + 1)}
                new_row = {c: new_data.get((r, c)) for c in range(1, new_max_c + 1)}
                
                is_deleted = (r <= old_max_r and r > new_max_r)
                is_new = (r > old_max_r and r <= new_max_r)
                
                process_row(r, old_row, new_row, is_new, is_deleted)
                        
    if progress_callback:
        progress_callback("Saving report file...", 0.95)
        
    # --- Apply Conditional Formatting for 'out' logic ---
    try:
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        
        for sht_name in out_wb.sheetnames:
            if "_Suggestions" in sht_name:
                sug_sheet = out_wb[sht_name]
                rule_sug = FormulaRule(formula=['OR(LOWER($F2)="out", LOWER($F2)="x")'], stopIfTrue=True, fill=yellow_fill)
                sug_sheet.conditional_formatting.add('A2:F10000', rule_sug)
                
            elif "_Out_Of_Stock" in sht_name:
                oos_sheet = out_wb[sht_name]
                sug_sheet_name = sht_name.replace("_Out_Of_Stock", "_Suggestions")
                if sug_sheet_name in out_wb.sheetnames:
                    # Formula: =OR(LOWER(IFERROR(VLOOKUP($A2, 'Sheet1_Suggestions'!$A:$F, 6, FALSE), ""))="out", ...)
                    rule_oos = FormulaRule(
                        formula=[f'OR(LOWER(IFERROR(VLOOKUP($A2, \'{sug_sheet_name}\'!$A:$F, 6, FALSE), ""))="out", LOWER(IFERROR(VLOOKUP($A2, \'{sug_sheet_name}\'!$A:$F, 6, FALSE), ""))="x")'],
                        stopIfTrue=True,
                        fill=yellow_fill
                    )
                    oos_sheet.conditional_formatting.add('A2:F10000', rule_oos)
    except Exception as e:
        print(f"Error applying conditional formatting: {e}")
        
    out_wb.save(output_path)
    
    if progress_callback:
        progress_callback("Comparison complete!", 1.0)
